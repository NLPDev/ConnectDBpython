import pandas
import pymysql
from openpyxl import Workbook

#Create Excel
wb=Workbook()
ws=wb.active

shw = wb.active
shw.title = 'Sheet'

# Create a connection object
databaseServerIP = "127.0.0.1"  # IP address of the MySQL database server
databaseUserName = "root"  # User name of the database server
databaseUserPassword = ""  # Password for the database user
newDatabaseName = "NewDatabase"  # Name of the database that is to be created

charSet = "utf8mb4"  # Character set
cusrorType = pymysql.cursors.DictCursor

connectionInstance = pymysql.connect(host=databaseServerIP, user=databaseUserName, password=databaseUserPassword,

                                     charset=charSet, cursorclass=cusrorType)

try:
    cursorInsatnce = connectionInstance.cursor()
    sqlStatement = "CREATE DATABASE " + newDatabaseName
    cursorInsatnce.execute(sqlStatement)
    sqlQuery = "SHOW DATABASES"
    cursorInsatnce.execute(sqlQuery)
    databaseList = cursorInsatnce.fetchall()

    # for datatbase in databaseList:
    #     print(datatbase)

except Exception as e:
    print("Exeception occured:{}".format(e))


df=pandas.read_csv('event-data-extract-v0.1.csv')#Read data from csv

flag=1
for rr in df:
    item=rr.split('\t')
    lt=len(item)
    if flag:
        tit=item
        break

db = pymysql.connect("localhost","root","","NewDatabase" )
cursor = db.cursor()

cursor.execute("DROP TABLE IF EXISTS EMPLOYEE")

sql = "CREATE TABLE EMPLOYEE ("
sqlInsert = "INSERT INTO EMPLOYEE ("

cntrow=1    #Excel Row, Column number
cntcolumn=1

flag=1

setlen=df.loc[0].iat[0]

setst=setlen.split('\t')

si=0
sl=len(setlen)

for tt in tit:
    aa=tt.split('\"')
    if len(aa)>1:
        aa=aa[1]
    else:
        aa=aa[0]

    aa=aa.replace(".", "_")

    if si<sl:
        st = setst[si]
        si=si+1

    cell = shw.cell(row=cntrow, column=cntcolumn)
    cell.value = aa
    cntcolumn = cntcolumn + 1

    if flag:
        flag=0
        sql=sql+aa+" CHAR(20)"
        sqlInsert=sqlInsert+aa

    else:
        if len(st)>20:
            sql = sql + ", " + aa + " CHAR(200)"
            sqlInsert = sqlInsert + ", " + aa
        else:
            sql=sql+", "+aa+" CHAR(20)"
            sqlInsert=sqlInsert+", "+aa


sql=sql+")"

cursor.execute(sql) #Create Table

sqlInsert=sqlInsert+") VALUES ("

flag=1
i=0
for rr in df:
    cntrow=cntrow+1
    cntcolumn=1
    rr=df.loc[i].iat[0]
    i=i+1

    item = rr.split('\t')
    sql = sqlInsert

    ff = 1

    for aa in item:
        aa = aa.split('\"')
        if len(aa) > 1:
            aa = aa[1]
        else:
            aa = aa[0]

        cell = shw.cell(row=cntrow, column=cntcolumn)
        cell.value = aa
        cntcolumn = cntcolumn + 1

        if ff:
            ff = 0
            sql = sql + "\""+aa+"\""

        else:
            sql = sql + ", \"" + aa + "\""


    for ii in range(lt-len(item)):
        sql=sql+", "+"\"  \""

    sql=sql+")"

    try:
        # Execute the SQL command
        cursor.execute(sql)
        # Commit your changes in the database
        db.commit()
    except:
        # Rollback in case there is any error
        db.rollback()



wb.save("res.xlsx")
# disconnect from server
db.close()

connectionInstance.close()

